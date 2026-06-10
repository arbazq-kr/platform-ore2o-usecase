"""
DEV NOTES: ARBAZ QURESHI (TTC4624)
codeExcel - Version 3 (Excel Based)

Matches incidents using two-layer category/sub_category matching.
Updates Excel file with parent-child linking results.

TWO LAYER MATCHING:
Layer 1 → normalized category     (broad match)
Layer 2 → normalized sub_category (precise match)

PARENT LOGIC:
If parent exists in group    → link unlinked incidents to it
If no parent exists in group → make oldest incident the parent
                               (oldest must not be child of another incident)
"""

import pandas as pd
import re
import os
from datetime import datetime
from openpyxl import load_workbook

# ══════════════════════════════════════════════════════════
#  CONFIG
#  Update EXCEL_FILE_PATH to point to your Excel file
# ══════════════════════════════════════════════════════════

#  Excel file path
# ──────────────────────────────────────────────────────────
EXCEL_FILE_PATH  = os.path.join(os.path.dirname(__file__), "incidents.xlsx")
EXCEL_SHEET_NAME = "incidents"

# ──────────────────────────────────────────────────────────
#  Output Excel file path
OUTPUT_FILE_PATH = os.path.join(os.path.dirname(__file__), "incidents_output.xlsx")

# ──────────────────────────────────────────────────────────
#  Column names in Excel
# ──────────────────────────────────────────────────────────
COL_NUMBER       = "number"
COL_SYS_ID       = "sys_id"
COL_STATE        = "state"
COL_CATEGORY     = "category"
COL_SUB_CATEGORY = "sub_category"
COL_SHORT_DESC   = "short_description"
COL_PARENT       = "parent_incident"
COL_CHILD_COUNT  = "child_incidents"
COL_CREATED_ON   = "sys_created_on"

# ──────────────────────────────────────────────────────────
#  Columns that will be ADDED by this script to output Excel
# ──────────────────────────────────────────────────────────
COL_NORM_CATEGORY     = "normalized_category"
COL_NORM_SUB_CATEGORY = "normalized_sub_category"
COL_SUGGESTED_PARENT  = "suggested_parent"
COL_ACTION            = "action"
COL_PROCESSED_AT      = "processed_at"


# ══════════════════════════════════════════════════════════
#  FUNCTION: NORMALIZE TEXT
#  Cleans and normalizes category / sub_category text
#  Used for consistent groupby matching
# ══════════════════════════════════════════════════════════

def normalize_text(text):
    """
    Normalizes text for consistent matching:
    1. Convert to lowercase
    2. Remove special characters
    3. Strip extra whitespace
    4. Strip leading/trailing spaces

    Args:
        text → raw category or sub_category string

    Returns:
        Normalized string or None if empty
    """
    if not text or str(text).strip() == "" or str(text).strip().lower() == "nan":
        return None

    # ─── Convert to lowercase ─────────────────────────────
    # "Network" → "network"
    # "SERVER"  → "server"
    text = str(text).lower()

    # ─── Remove special characters ────────────────────────
    # Keep only letters, numbers, spaces
    text = re.sub(r'[^a-z0-9\s]', ' ', text)

    # ─── Remove extra whitespace ──────────────────────────
    # "server  down" → "server down"
    text = re.sub(r'\s+', ' ', text).strip()

    return text if text else None


# ══════════════════════════════════════════════════════════
#  FUNCTION: GET PARENT SYS ID
#  Extracts parent sys_id from parent_incident field
# ══════════════════════════════════════════════════════════

def get_parent_sys_id(parent_incident_field):
    """
    Extracts and returns parent sys_id string from
    parent_incident field regardless of its type.
    Returns None if no valid parent found.

    Args:
        parent_incident_field → value from parent_incident column

    Returns:
        sys_id string or None
    """
    if parent_incident_field is None:
        return None
    if str(parent_incident_field).strip() == "":
        return None
    if str(parent_incident_field).strip().lower() == "nan":
        return None
    return str(parent_incident_field).strip()


# ══════════════════════════════════════════════════════════
#  FUNCTION: ASK USER PERMISSION
#  Asks Y/N question and returns True/False
# ══════════════════════════════════════════════════════════

def ask_permission(question):
    """
    Ask user Y/N question.
    Returns True if Yes, False if No.
    Keeps asking until valid input given.

    Args:
        question → string question to ask

    Returns:
        True if Y, False if N
    """
    while True:
        answer = input(f"\n   {question} (Y/N): ").strip().upper()

        if answer == 'Y':
            return True
        elif answer == 'N':
            return False
        else:
            print(f"   ⚠️  Invalid input '{answer}'. Please enter Y or N only.")


# ══════════════════════════════════════════════════════════
#  FUNCTION: FETCH DATA
#  Reads incident data from Excel file
#  (Will be replaced with ServiceNow API call later)
# ══════════════════════════════════════════════════════════

def fetch_data():
    """
    Reads incident data from Excel file.
    Later this function will be updated to
    call ServiceNow API instead.

    Returns:
        pandas DataFrame with incident data
        or None if file not found / error
    """
    print(f"\n{'='*65}")
    print(f"   📊 INCIDENT PARENT-CHILD LINKER (Excel Mode)")
    print(f"{'='*65}")
    print(f"\n🔄 Reading incidents from Excel...")
    print(f"   File  : {EXCEL_FILE_PATH}")
    print(f"   Sheet : {EXCEL_SHEET_NAME}")

    # ─── Check if file exists ─────────────────────────────
    if not os.path.exists(EXCEL_FILE_PATH):
        print(f"\n❌ Excel file not found!")
        print(f"   Expected at: {EXCEL_FILE_PATH}")
        print(f"   Please create incidents.xlsx in same folder as this script")
        return None

    # ─── Read Excel file ──────────────────────────────────
    df = pd.read_excel(
        EXCEL_FILE_PATH,
        sheet_name = EXCEL_SHEET_NAME,
        dtype      = str         # read all columns as string
                                 # avoids type conversion issues
    )

    # ─── Strip whitespace from all string columns ─────────
    df = df.apply(lambda col: col.str.strip() if col.dtype == "object" else col)

    # ─── Replace empty strings with None ──────────────────
    df = df.replace("", None)
    df = df.replace("nan", None)

    print(f"✅ Read {len(df)} incidents from Excel")

    return df


# ══════════════════════════════════════════════════════════
#  FUNCTION: NORMALIZE DATA
#  Adds normalized_category and normalized_sub_category
#  columns to DataFrame for consistent matching
# ══════════════════════════════════════════════════════════

def normalize_data(df):
    """
    Adds normalized text columns to DataFrame.
    Normalizes category and sub_category for matching.

    Normalization:
    → lowercase
    → remove special characters
    → strip extra whitespace

    Args:
        df → raw DataFrame from fetch_data()

    Returns:
        DataFrame with added normalized columns
    """
    print(f"\n{'='*65}")
    print(f"🔤 Normalizing Category and Sub-Category Text...")
    print(f"{'='*65}")

    # ─── Normalize category ───────────────────────────────
    # Example: "Network Issues" → "network issues"
    df[COL_NORM_CATEGORY] = df[COL_CATEGORY].apply(normalize_text)

    # ─── Normalize sub_category ───────────────────────────
    # Example: "Server Down!" → "server down"
    df[COL_NORM_SUB_CATEGORY] = df[COL_SUB_CATEGORY].apply(normalize_text)

    # ─── Print sample of normalization ────────────────────
    print(f"\n   📋 Normalization Sample:")
    print(f"   {'Number':<12} {'Category':<20} {'Normalized':<20} {'SubCat':<20} {'Normalized'}")
    print(f"   {'-'*95}")

    for _, row in df.head(10).iterrows():
        print(
            f"   {str(row[COL_NUMBER]):<12} "
            f"{str(row[COL_CATEGORY]):<20} "
            f"{str(row[COL_NORM_CATEGORY]):<20} "
            f"{str(row[COL_SUB_CATEGORY]):<20} "
            f"{str(row[COL_NORM_SUB_CATEGORY])}"
        )

    # ─── Add output columns with default values ───────────
    # These will be filled in during matching
    df[COL_SUGGESTED_PARENT] = None
    df[COL_ACTION]           = "no_action"
    df[COL_PROCESSED_AT]     = None

    print(f"\n✅ Normalization complete!")

    return df


# ══════════════════════════════════════════════════════════
#  FUNCTION: FIND SIMILAR INCIDENTS
#  Two-layer matching using category and sub_category
#  Layer 1 → normalized_category     (broad match)
#  Layer 2 → normalized_sub_category (precise match)
# ══════════════════════════════════════════════════════════

def find_similar_incidents(df):
    """
    Finds similar incidents using two-layer groupby matching.

    Layer 1 → groups by normalized_category
              All incidents with same category
              are considered broadly similar

    Layer 2 → within each Layer 1 group,
              sub-groups by normalized_sub_category
              Only incidents with same category AND
              same sub_category are truly similar

    Args:
        df → normalized DataFrame

    Returns:
        List of sub-groups
        Each sub-group is a list of DataFrame indices
        Example: [[0,3,7], [1,5], [2,6,8,9]]
    """
    print(f"\n{'='*65}")
    print(f"🔍 Finding Similar Incidents (Two-Layer Matching)...")
    print(f"   Layer 1 → normalized_category")
    print(f"   Layer 2 → normalized_sub_category")
    print(f"{'='*65}")

    all_subgroups = []

    # ══════════════════════════════════════════════════════
    #  LAYER 1: Group by normalized_category
    #  Find all incidents with same category
    # ══════════════════════════════════════════════════════

    # ─── Filter out rows with no category ─────────────────
    # Cannot group incidents with no category
    df_with_cat    = df[df[COL_NORM_CATEGORY].notna()].copy()
    df_without_cat = df[df[COL_NORM_CATEGORY].isna()].copy()

    if not df_without_cat.empty:
        print(f"\n   ⚠️  {len(df_without_cat)} incident(s) have no category → skipped:")
        for _, row in df_without_cat.iterrows():
            print(f"      → {row[COL_NUMBER]}")

    # ─── Group by normalized_category ─────────────────────
    # Only keep groups with 2+ incidents
    layer1_similar = df_with_cat.groupby(COL_NORM_CATEGORY).filter(
        lambda x: len(x) > 1
    )

    if layer1_similar.empty:
        print(f"\n✅ No similar incidents found in Layer 1. Nothing to link!")
        return []

    total_l1 = layer1_similar.groupby(COL_NORM_CATEGORY).ngroups
    current_l1 = 0

    print(f"\n⚠️  Found {total_l1} broad category group(s) from Layer 1!")

    for category, l1_group in layer1_similar.groupby(COL_NORM_CATEGORY):

        current_l1 += 1

        print(f"\n{'-'*65}")
        print(f"🔵  Layer 1 Group {current_l1}/{total_l1}")
        print(f"   Category   : '{category}'")
        print(f"   Total INCs : {len(l1_group)}")

        # ══════════════════════════════════════════════════
        #  LAYER 2: Sub-group by normalized_sub_category
        #  Within each category group, find incidents
        #  with same sub_category
        # ══════════════════════════════════════════════════

        # ─── Filter out rows with no sub_category ─────────
        l1_with_sub    = l1_group[l1_group[COL_NORM_SUB_CATEGORY].notna()].copy()
        l1_without_sub = l1_group[l1_group[COL_NORM_SUB_CATEGORY].isna()].copy()

        if not l1_without_sub.empty:
            print(f"\n   ⚠️  {len(l1_without_sub)} incident(s) have no sub_category → skipped:")
            for _, row in l1_without_sub.iterrows():
                print(f"      → {row[COL_NUMBER]}")

        # ─── Group by normalized_sub_category ─────────────
        # Only keep sub-groups with 2+ incidents
        layer2_similar = l1_with_sub.groupby(COL_NORM_SUB_CATEGORY).filter(
            lambda x: len(x) > 1
        )

        if layer2_similar.empty:
            print(f"\n   ✅ No exact sub-category matches in Layer 2")
            print(f"      All incidents have different sub-categories")
            print(f"      No linking needed for this category group")
            continue

        total_l2   = layer2_similar.groupby(COL_NORM_SUB_CATEGORY).ngroups
        current_l2 = 0

        print(f"\n   ✅ Found {total_l2} sub-category sub-group(s) in Layer 2!")

        for sub_category, l2_group in layer2_similar.groupby(COL_NORM_SUB_CATEGORY):

            current_l2 += 1

            print(f"\n   📌 Layer 2 Sub-Group {current_l2}/{total_l2}")
            print(f"      Sub-Category : '{sub_category}'")
            print(f"      Total INCs   : {len(l2_group)}")

            # ─── Add this sub-group to results ────────────
            # Store as list of DataFrame indices
            all_subgroups.append(l2_group.index.tolist())

    print(f"\n{'='*65}")
    print(f"✅ Total sub-groups found for linking : {len(all_subgroups)}")

    return all_subgroups


# ══════════════════════════════════════════════════════════
#  FUNCTION: DETERMINE PARENT FOR GROUP
#  Finds or creates parent for a group of similar incidents
# ══════════════════════════════════════════════════════════

def determine_parent(group_sorted):
    """
    Determines which incident is or becomes the parent.

    Logic:
    1. Check if any incident in group has parent_incident
       pointing to ANOTHER incident IN THE SAME GROUP
       → If yes → that is the existing valid parent
       → Link all unlinked incidents to it

    2. If no in-group parent found:
       → Find oldest incident that has NO parent at all
       → Make it the parent
       → Link all others to it

    Args:
        group_sorted → DataFrame sorted oldest first

    Returns:
        tuple:
            parent_sys_id  → sys_id of chosen parent
            parent_number  → number of chosen parent
            is_existing    → True if parent already existed
                             False if we are making new parent
    """
    # ─── Collect all sys_ids in this group ────────────────
    group_sys_ids = set(group_sorted[COL_SYS_ID].tolist())

    # ══════════════════════════════════════════════════════
    #  CHECK 1: Does any incident have parent pointing
    #           to another incident IN THIS GROUP?
    # ══════════════════════════════════════════════════════

    for _, row in group_sorted.iterrows():
        parent_field = get_parent_sys_id(row[COL_PARENT])

        if parent_field is not None:
            if parent_field in group_sys_ids:
                # ✅ Found valid in-group parent
                # The incident whose sys_id is parent_field
                # is the parent
                parent_row = group_sorted[
                    group_sorted[COL_SYS_ID] == parent_field
                ].iloc[0]

                return (
                    parent_field,
                    parent_row[COL_NUMBER],
                    True            # is_existing = True
                )

    # ══════════════════════════════════════════════════════
    #  CHECK 2: No in-group parent found
    #           Find oldest incident with NO parent at all
    #           Make it the parent
    # ══════════════════════════════════════════════════════

    for _, row in group_sorted.iterrows():
        parent_field = get_parent_sys_id(row[COL_PARENT])

        if parent_field is None:
            # ✅ This incident has no parent
            # It is eligible to be the parent
            return (
                row[COL_SYS_ID],
                row[COL_NUMBER],
                False           # is_existing = False (new parent)
            )

    # ══════════════════════════════════════════════════════
    #  CHECK 3: All incidents have external parents
    #           Cannot determine parent
    #           Skip this group
    # ══════════════════════════════════════════════════════

    return None, None, None


# ══════════════════════════════════════════════════════════
#  FUNCTION: PROCESS GROUP
#  Processes one sub-group of similar incidents
#  Shows plan, asks permission, records actions
# ══════════════════════════════════════════════════════════

def process_group(df, group_indices, current_group, total_groups):
    """
    Processes one sub-group of similar incidents.
    1. Sorts group oldest first
    2. Prints group info
    3. Determines parent
    4. Shows linking plan
    5. Asks user permission
    6. Records actions in DataFrame

    Args:
        df            → full DataFrame
        group_indices → list of DataFrame indices in this group
        current_group → current group number (for display)
        total_groups  → total groups count (for display)

    Returns:
        df → updated DataFrame with actions recorded
    """
    # ─── Get group rows from DataFrame ────────────────────
    group = df.loc[group_indices].copy()

    # ─── Convert sys_created_on to datetime for sorting ───
    group[COL_CREATED_ON] = pd.to_datetime(
        group[COL_CREATED_ON],
        errors = 'coerce'       # invalid dates become NaT
    )

    # ─── Sort oldest first ────────────────────────────────
    group_sorted = group.sort_values(COL_CREATED_ON, ascending=True)

    oldest        = group_sorted.iloc[0]
    oldest_number = oldest[COL_NUMBER]

    # ─── Print Group Info ─────────────────────────────────
    print(f"\n{'═'*65}")
    print(f"📌 Sub-Group {current_group}/{total_groups}")
    print(f"   Category     : {oldest[COL_CATEGORY]}")
    print(f"   Sub-Category : {oldest[COL_SUB_CATEGORY]}")
    print(f"   Total Count  : {len(group)} incidents")
    print(f"\n   {'Number':<15} {'State':<8} {'Created On':<25} {'Parent':<20} {'Role'}")
    print(f"   {'-'*85}")

    for _, row in group_sorted.iterrows():
        role         = "⭐ Oldest" if row[COL_NUMBER] == oldest_number else "👶 Newer"
        parent_val   = str(row[COL_PARENT]) if get_parent_sys_id(row[COL_PARENT]) else "None"
        print(
            f"   {str(row[COL_NUMBER]):<15} "
            f"{str(row[COL_STATE]):<8} "
            f"{str(row[COL_CREATED_ON]):<25} "
            f"{parent_val:<20} "
            f"{role}"
        )

    # ─── Determine parent ─────────────────────────────────
    parent_sys_id, parent_number, is_existing = determine_parent(group_sorted)

    # ─── Handle case where no parent can be determined ────
    if parent_sys_id is None:
        print(f"\n   ❌ All incidents have external parents!")
        print(f"      Cannot determine parent. Skipping group...")
        print(f"{'═'*65}")
        return df

    # ─── Show linking plan ────────────────────────────────
    print(f"\n   📋 LINKING PLAN:")

    if is_existing:
        # ─── Existing in-group parent found ───────────────
        print(f"   ✅ Existing parent found  : {parent_number}")
        print(f"   Will link UNLINKED incidents to this parent")
    else:
        # ─── Making oldest free incident the parent ────────
        print(f"   👑 No existing parent found")
        print(f"   Will make {parent_number} the PARENT (oldest with no parent)")

    # ─── Show what will happen to each incident ───────────
    link_count = 0
    skip_count = 0

    for _, row in group_sorted.iterrows():
        inc_number    = row[COL_NUMBER]
        inc_sys_id    = row[COL_SYS_ID]
        current_parent = get_parent_sys_id(row[COL_PARENT])

        # Skip if this IS the parent
        if inc_sys_id == parent_sys_id:
            print(f"   👑 {inc_number} → IS the parent (skip)")
            skip_count += 1
            continue

        # Skip if already linked to this parent
        if current_parent == parent_sys_id:
            print(f"   ⏭️  {inc_number} → already linked to parent (skip)")
            skip_count += 1
            continue

        # Will be linked
        print(f"   → Will link : {inc_number} as child of {parent_number}")
        link_count += 1

    # ─── If nothing to link → skip ────────────────────────
    if link_count == 0:
        print(f"\n   ✅ All incidents already linked! Nothing to do.")
        print(f"{'═'*65}")
        return df

    # ─── Ask user permission ──────────────────────────────
    user_approved = ask_permission(
        f"Do you want to link these {len(group)} incidents?"
    )

    if not user_approved:
        print(f"\n   🚫 SKIPPED by user")
        print(f"{'═'*65}")
        return df

    print(f"\n   ✅ User approved! Recording linking actions...")
    print(f"   {'-'*50}")

    # ─── Record actions in DataFrame ──────────────────────
    linked  = 0
    skipped = 0
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    for _, row in group_sorted.iterrows():
        inc_sys_id     = row[COL_SYS_ID]
        inc_number     = row[COL_NUMBER]
        current_parent = get_parent_sys_id(row[COL_PARENT])

        # ─── Skip if IS the parent ────────────────────────
        if inc_sys_id == parent_sys_id:
            # Mark as parent in output
            df.loc[df[COL_SYS_ID] == inc_sys_id, COL_ACTION]          = "is_parent"
            df.loc[df[COL_SYS_ID] == inc_sys_id, COL_SUGGESTED_PARENT] = ""
            df.loc[df[COL_SYS_ID] == inc_sys_id, COL_PROCESSED_AT]    = now_str
            print(f"   👑 {inc_number} → marked as PARENT")
            skipped += 1
            continue

        # ─── Skip if already linked to this parent ────────
        if current_parent == parent_sys_id:
            df.loc[df[COL_SYS_ID] == inc_sys_id, COL_ACTION]          = "already_linked"
            df.loc[df[COL_SYS_ID] == inc_sys_id, COL_SUGGESTED_PARENT] = parent_number
            df.loc[df[COL_SYS_ID] == inc_sys_id, COL_PROCESSED_AT]    = now_str
            print(f"   ⏭️  {inc_number} → already linked (skip)")
            skipped += 1
            continue

        # ─── Link this incident to parent ─────────────────
        # Record suggested parent in output
        # (actual API call will be added later)
        df.loc[df[COL_SYS_ID] == inc_sys_id, COL_ACTION]          = "link_to_parent"
        df.loc[df[COL_SYS_ID] == inc_sys_id, COL_SUGGESTED_PARENT] = parent_number
        df.loc[df[COL_SYS_ID] == inc_sys_id, COL_PROCESSED_AT]    = now_str
        print(f"   ✅ {inc_number} → recorded as child of {parent_number}")
        linked += 1

    # ─── Summary for this group ───────────────────────────
    print(f"\n   📊 Group Summary:")
    print(f"   👑 Parent  : {parent_number}")
    print(f"   ✅ Linked  : {linked}")
    print(f"   ⏭️  Skipped : {skipped}")
    print(f"{'═'*65}")

    return df


# ══════════════════════════════════════════════════════════
#  FUNCTION: MATCH AND PROCESS INCIDENTS
#  Orchestrates two-layer matching and processes each group
# ══════════════════════════════════════════════════════════

def match_and_process(df):
    """
    Orchestrates the full matching and processing flow.
    1. Calls find_similar_incidents() to get sub-groups
    2. Loops through each sub-group
    3. Calls process_group() for each sub-group

    Args:
        df → normalized DataFrame

    Returns:
        df → updated DataFrame with all actions recorded
    """
    # ─── Find similar incident groups ─────────────────────
    all_subgroups = find_similar_incidents(df)

    if not all_subgroups:
        print(f"\n✅ No similar incidents found. Nothing to link!")
        return df

    total_groups   = len(all_subgroups)
    current_group  = 0

    print(f"\n{'='*65}")
    print(f"🔗 Processing {total_groups} similar sub-group(s)...")
    print(f"{'='*65}")

    # ─── Process each sub-group ───────────────────────────
    for group_indices in all_subgroups:

        current_group += 1

        # ─── Process this sub-group ───────────────────────
        df = process_group(
            df            = df,
            group_indices = group_indices,
            current_group = current_group,
            total_groups  = total_groups
        )

    return df


# ══════════════════════════════════════════════════════════
#  FUNCTION: SAVE OUTPUT
#  Saves updated DataFrame to output Excel file
#  (Will be replaced with ServiceNow PATCH API later)
# ══════════════════════════════════════════════════════════

def save_output(df):
    """
    Saves updated DataFrame to output Excel file.
    Adds color coding to action column for easy reading.
    Later this function will call ServiceNow PATCH API
    instead of saving to Excel.

    Color coding in output Excel:
    Green  → link_to_parent  (will be linked)
    Yellow → already_linked  (already correct)
    Blue   → is_parent       (this is the parent)
    White  → no_action       (not in any group)

    Args:
        df → updated DataFrame with actions recorded
    """
    print(f"\n{'='*65}")
    print(f"💾 Saving Output to Excel...")
    print(f"   File : {OUTPUT_FILE_PATH}")
    print(f"{'='*65}")

    # ─── Save to Excel ────────────────────────────────────
    df.to_excel(
        OUTPUT_FILE_PATH,
        sheet_name = "results",
        index      = False
    )

    # ─── Add color coding using openpyxl ──────────────────
    # Load the saved file to add colors
    wb = load_workbook(OUTPUT_FILE_PATH)
    ws = wb["results"]

    # ─── Import openpyxl styles ───────────────────────────
    from openpyxl.styles import PatternFill, Font

    # ─── Define colors ────────────────────────────────────
    # Green  → link_to_parent
    # Yellow → already_linked
    # Blue   → is_parent
    # White  → no_action
    color_map = {
        "link_to_parent" : PatternFill("solid", fgColor="C6EFCE"),  # green
        "already_linked" : PatternFill("solid", fgColor="FFEB9C"),  # yellow
        "is_parent"      : PatternFill("solid", fgColor="BDD7EE"),  # blue
        "no_action"      : PatternFill("solid", fgColor="FFFFFF"),  # white
    }

    # ─── Find action column index ─────────────────────────
    # openpyxl is 1-indexed
    action_col_idx = None
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == COL_ACTION:
            action_col_idx = col_idx
            break

    # ─── Apply colors row by row ──────────────────────────
    if action_col_idx:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            action_cell = row[action_col_idx - 1]
            action_val  = str(action_cell.value) if action_cell.value else "no_action"
            fill        = color_map.get(action_val, color_map["no_action"])

            # Apply color to entire row
            for cell in row:
                cell.fill = fill

    # ─── Make header row bold ─────────────────────────────
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # ─── Auto-fit column widths ───────────────────────────
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    # ─── Save workbook with colors ────────────────────────
    wb.save(OUTPUT_FILE_PATH)

    # ─── Print summary of actions ─────────────────────────
    print(f"\n   📊 Output Summary:")
    print(f"   {'Action':<20} {'Count'}")
    print(f"   {'-'*30}")

    action_counts = df[COL_ACTION].value_counts()
    for action, count in action_counts.items():
        print(f"   {str(action):<20} {count}")

    print(f"\n✅ Output saved successfully!")
    print(f"   Open {OUTPUT_FILE_PATH} to see results")


# ══════════════════════════════════════════════════════════
#  MAIN FUNCTION
# ══════════════════════════════════════════════════════════

def main():
    """
    Main entry point.
    Orchestrates all steps:
    1. fetch_data()          → read from Excel
    2. normalize_data()      → clean category/sub_category
    3. match_and_process()   → two-layer matching + linking
    4. save_output()         → write results to Excel
    """

    # ─── Print start time ─────────────────────────────────
    start_time = datetime.now()
    print(f"🕐 Start Time : {start_time.strftime('%Y-%m-%d %H:%M:%S')}")

    # ─── STEP 1: Fetch data from Excel ────────────────────
    # Later: replace with ServiceNow API call
    df = fetch_data()

    if df is None or df.empty:
        print(f"\n✅ No data found. Nothing to process!")
        return

    # ─── STEP 2: Normalize category and sub_category ──────
    df = normalize_data(df)

    # ─── STEP 3: Match and process similar incidents ───────
    # Two-layer matching: category → sub_category
    df = match_and_process(df)

    # ─── STEP 4: Save output to Excel ─────────────────────
    # Later: replace with ServiceNow PATCH API call
    save_output(df)

    # ─── Print end time and duration ──────────────────────
    end_time     = datetime.now()
    elapsed_time = end_time - start_time

    print(f"\n{'='*65}")
    print(f"   ✅ ALL DONE!")
    print(f"   → Data read from Excel")
    print(f"   → Category and Sub-Category normalized")
    print(f"   → Similar incidents identified (2-layer match)")
    print(f"   → Parent-Child actions recorded")
    print(f"   → Results saved to output Excel")
    print(f"{'='*65}")
    print(f"🕐 End Time   : {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"⏱️  Total Time : {str(elapsed_time).split('.')[0]}")
    print(f"{'='*65}")


# ══════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════

if __name__ == "__main__":
    main()